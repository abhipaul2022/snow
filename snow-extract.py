import os as os
import pandas as pd
from openpyxl import load_workbook
import openpyxl as pxl
from datetime import date,datetime,timedelta
from openpyxl.styles import Font,Alignment

   
print("Generating Daily Tickets Report")
print('*************************')
print("Reading Parameter File......,")    
df=pd.read_excel("Parameters File.xlsx",sheet_name='Daily Tickets')
df0=pd.read_excel("Parameters File.xlsx",sheet_name='FPaths')
xpath=os.getcwd()+df0["Path"][0]


for f in range(len(df)):
  if df['Completed(Y/N)'][f]=='N':
    filename=xpath+'\\'+df["File Name"][f]+'.xlsx'
    print("Processing ",filename)  
    df1=pd.read_excel(filename,sheet_name=' Incident')
    df1['Dummy']= pd.to_datetime(df1['Updated']).dt.date
    df2=df1
    df3=df1
    df4=df1
    df5=df1
    df6=df1
    df7=df1
    df8=df1
    df9=df1

    df2=df2[(df2["Ticket type"]== 'Incident') & (df2["GD Resource"]!= 'Not GD') & (pd.isnull(df2["Close code"]) )]
    df3=df3[(df3["Ticket type"]== 'Service Request') & (df3["GD Resource"]!= 'Not GD') & (df3["State"] == 'Work In Progress') & (df3["Dummy"] != df["Date"][f]) & (df3["Dummy"] != df["Date"][f].date() - timedelta(days=1))]
    df4=df4[(df4["Ticket type"]== 'Service Request') & (df4["GD Resource"]!= 'Not GD') &  (df4["State"] == 'On Hold') & (df4["On hold reason"] == 'Awaiting User input') & (df4["Dummy"] != df["Date"][f]) & (df4["Dummy"] != df["Date"][f].date() - timedelta(days= 1))]
    df5=df5[(df5["Ticket type"]== 'Service Request') & (df5["GD Resource"]!= 'Not GD') &  (df5["State"] == 'On Hold') & (df5["On hold reason"] == 'Customer Testing') & (df5["Dummy"] != df["Date"][f]) & (df5["Dummy"] != df["Date"][f].date() - timedelta(days= 1))]
    df6=df6[(df6["Ticket type"]== 'Service Request') & (df6["GD Resource"]!= 'Not GD') &  (df6["State"] == 'On Hold') & (df6["On hold reason"] == 'Monitoring') & (df6["Dummy"] != df["Date"][f]) & (df6["Dummy"] != df["Date"][f].date() - timedelta(days= 1))]
    df7=df7[(df7["Ticket type"]== 'Service Request') & (df7["GD Resource"]!= 'Not GD') & (df7["State"] == 'On Hold') & (df7["On hold reason"] == 'Non IBM 3rd Party Engagement') & (df7["Dummy"] != df["Date"][f]) & (df7["Dummy"] != df["Date"][f].date() - timedelta(days= 1)) ]
    df8=df8[(df8["Ticket type"]== 'Service Request') & (df8["GD Resource"]!= 'Not GD')  & (df8["State"] == 'On Hold') & (df8["On hold reason"] == 'Work not yet due') & (df8["Dummy"] != df["Date"][f]) & (df8["Dummy"] != df["Date"][f].date() - timedelta(days= 1))]
    df9=df9[(df9["Ticket type"]== 'Service Request') & (df9["GD Resource"]!= 'Not GD') & (df9["State"] == 'On Hold') & ((df9["On hold reason"] == 'Customer unavailable 1st Attempt')|(df9["On hold reason"] == 'Customer unavailable 2nd Attempt'))]

    a1=len(df2)
    a2=len(df3)
    a3=len(df4)
    a4=len(df5)
    a5=len(df6)
    a6=len(df7)
    a7=len(df8)
    a8=len(df9)

    df2.drop(['Assigned on','Close code','Close notes','Closed','Closed by','Created','Created by','Escalated','Escalated by','SLA Class','Escalated count','Escalation','State','Resolved','Resolve time','Reported by','Reopen count','SLA hold','FIrst_Assigned','First Assigned to Service Desk','Reassignment count','First Assigned to Resolver Group','First Assigned to Osprey-Resolver','First Assigned to Osprey-RG','Region','On hold reason','OpCo','Days Open','Aging','Week Created','Week Resolved','Week Closed','GD Resource','Type of Resource','Year Closed','Year Resolved','year Created','Day Res','Month Res','Day Crted','Month Crted','Team','Dummy'],inplace=True,axis=1)
    df3.drop(['Assigned on','Close code','Close notes','Closed','Closed by','Created','Created by','Escalated','Escalated by','SLA Class','Escalated count','Escalation','State','Resolved','Resolve time','Reported by','Reopen count','SLA hold','FIrst_Assigned','First Assigned to Service Desk','Reassignment count','First Assigned to Resolver Group','First Assigned to Osprey-Resolver','First Assigned to Osprey-RG','Region','On hold reason','OpCo','Days Open','Aging','Week Created','Week Resolved','Week Closed','GD Resource','Type of Resource','Year Closed','Year Resolved','year Created','Day Res','Month Res','Day Crted','Month Crted','Team','Dummy'],inplace=True,axis=1)
    df4.drop(['Assigned on','Close code','Close notes','Closed','Closed by','Created','Created by','Escalated','Escalated by','SLA Class','Escalated count','Escalation','State','Resolved','Resolve time','Reported by','Reopen count','SLA hold','FIrst_Assigned','First Assigned to Service Desk','Reassignment count','First Assigned to Resolver Group','First Assigned to Osprey-Resolver','First Assigned to Osprey-RG','Region','On hold reason','OpCo','Days Open','Aging','Week Created','Week Resolved','Week Closed','GD Resource','Type of Resource','Year Closed','Year Resolved','year Created','Day Res','Month Res','Day Crted','Month Crted','Team','Dummy'],inplace=True,axis=1)
    df5.drop(['Assigned on','Close code','Close notes','Closed','Closed by','Created','Created by','Escalated','Escalated by','SLA Class','Escalated count','Escalation','State','Resolved','Resolve time','Reported by','Reopen count','SLA hold','FIrst_Assigned','First Assigned to Service Desk','Reassignment count','First Assigned to Resolver Group','First Assigned to Osprey-Resolver','First Assigned to Osprey-RG','Region','On hold reason','OpCo','Days Open','Aging','Week Created','Week Resolved','Week Closed','GD Resource','Type of Resource','Year Closed','Year Resolved','year Created','Day Res','Month Res','Day Crted','Month Crted','Team','Dummy'],inplace=True,axis=1)
    df6.drop(['Assigned on','Close code','Close notes','Closed','Closed by','Created','Created by','Escalated','Escalated by','SLA Class','Escalated count','Escalation','State','Resolved','Resolve time','Reported by','Reopen count','SLA hold','FIrst_Assigned','First Assigned to Service Desk','Reassignment count','First Assigned to Resolver Group','First Assigned to Osprey-Resolver','First Assigned to Osprey-RG','Region','On hold reason','OpCo','Days Open','Aging','Week Created','Week Resolved','Week Closed','GD Resource','Type of Resource','Year Closed','Year Resolved','year Created','Day Res','Month Res','Day Crted','Month Crted','Team','Dummy'],inplace=True,axis=1)
    df7.drop(['Assigned on','Close code','Close notes','Closed','Closed by','Created','Created by','Escalated','Escalated by','SLA Class','Escalated count','Escalation','State','Resolved','Resolve time','Reported by','Reopen count','SLA hold','FIrst_Assigned','First Assigned to Service Desk','Reassignment count','First Assigned to Resolver Group','First Assigned to Osprey-Resolver','First Assigned to Osprey-RG','Region','On hold reason','OpCo','Days Open','Aging','Week Created','Week Resolved','Week Closed','GD Resource','Type of Resource','Year Closed','Year Resolved','year Created','Day Res','Month Res','Day Crted','Month Crted','Team','Dummy'],inplace=True,axis=1)
    df8.drop(['Assigned on','Close code','Close notes','Closed','Closed by','Created','Created by','Escalated','Escalated by','SLA Class','Escalated count','Escalation','State','Resolved','Resolve time','Reported by','Reopen count','SLA hold','FIrst_Assigned','First Assigned to Service Desk','Reassignment count','First Assigned to Resolver Group','First Assigned to Osprey-Resolver','First Assigned to Osprey-RG','Region','On hold reason','OpCo','Days Open','Aging','Week Created','Week Resolved','Week Closed','GD Resource','Type of Resource','Year Closed','Year Resolved','year Created','Day Res','Month Res','Day Crted','Month Crted','Team','Dummy'],inplace=True,axis=1)
    df9.drop(['Assigned on','Close code','Close notes','Closed','Closed by','Created','Created by','Escalated','Escalated by','SLA Class','Escalated count','Escalation','State','Resolved','Resolve time','Reported by','Reopen count','SLA hold','FIrst_Assigned','First Assigned to Service Desk','Reassignment count','First Assigned to Resolver Group','First Assigned to Osprey-Resolver','First Assigned to Osprey-RG','Region','On hold reason','OpCo','Days Open','Aging','Week Created','Week Resolved','Week Closed','GD Resource','Type of Resource','Year Closed','Year Resolved','year Created','Day Res','Month Res','Day Crted','Month Crted','Team','Dummy'],inplace=True,axis=1)

    df2=df2.append(df3)
    df2=df2.append(df4)
    df2=df2.append(df5)
    df2=df2.append(df6)
    df2=df2.append(df7)
    df2=df2.append(df8)
    df2=df2.append(df9)


    excel_book = pxl.load_workbook(filename)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
         writer.book = excel_book
         writer.sheets = {worksheet.title: worksheet for worksheet in excel_book.worksheets}
         df2.to_excel(writer,sheet_name="SheetNew",index=False)
         writer.save()

    wb = load_workbook('Parameters File.xlsx')    
    wsh1 = wb.worksheets[0]        
    v1=wsh1.cell(row=f+2,column=3)
    v1.value='Y'
    wb.save('Parameters File.xlsx')

    wb1 = load_workbook(filename)
    ws1 = wb1.worksheets[-1]

    ws1.insert_rows(1)
    ws1['A1']='Incidents'

    for i in range(ws1.max_column):
      ws1.cell(row=2, column=i+1).alignment = Alignment(horizontal='left')
            
    ws1.insert_rows(a1+3)
    ws1.insert_rows(a1+4)
    ws1['A'+str(a1+4)]='WIP > 2 days'
    ws1.insert_rows(a1+5)

    for i in range(ws1.max_column):
      ws1.cell(row=a1+5,column=i+1).value = ws1.cell(row=2,column=i+1).value
      ws1.cell(row=a1+5,column=i+1).font=Font(bold=True)
            
    ws1.insert_rows(a1+5+a2+1)
    ws1.insert_rows(a1+5+a2+2)
    ws1['A'+str(a1+5+a2+2)]='Awaiting User Input'
    ws1.insert_rows(a1+5+a2+3)

    for i in range(ws1.max_column):
      ws1.cell(row=a1+5+a2+3,column=i+1).value = ws1.cell(row=2,column=i+1).value
      ws1.cell(row=a1+5+a2+3,column=i+1).font=Font(bold=True)

    ws1.insert_rows(a1+5+a2+3+a3+1)
    ws1.insert_rows(a1+5+a2+3+a3+2)
    ws1['A'+str(a1+5+a2+3+a3+2)]='Customer Testing'
    ws1.insert_rows(a1+5+a2+3+a3+3)

    for i in range(ws1.max_column):
      ws1.cell(row=a1+5+a2+3+a3+3,column=i+1).value = ws1.cell(row=2,column=i+1).value
      ws1.cell(row=a1+5+a2+3+a3+3,column=i+1).font=Font(bold=True)

    ws1.insert_rows(a1+5+a2+3+a3+3+a4+1)
    ws1.insert_rows(a1+5+a2+3+a3+3+a4+2)
    ws1['A'+str(a1+5+a2+3+a3+3+a4+2)]='Monitoring'
    ws1.insert_rows(a1+5+a2+3+a3+3+a4+3)

    for i in range(ws1.max_column):
      ws1.cell(row=a1+5+a2+3+a3+3+a4+3,column=i+1).value = ws1.cell(row=2,column=i+1).value
      ws1.cell(row=a1+5+a2+3+a3+3+a4+3,column=i+1).font=Font(bold=True)

    ws1.insert_rows(a1+5+a2+3+a3+3+a4+3+a5+1)
    ws1.insert_rows(a1+5+a2+3+a3+3+a4+3+a5+2)
    ws1['A'+str(a1+5+a2+3+a3+3+a4+3+a5+2)]='Non IBM 3rd Party'
    ws1.insert_rows(a1+5+a2+3+a3+3+a4+3+a5+3)

    for i in range(ws1.max_column):
      ws1.cell(row=a1+5+a2+3+a3+3+a4+3+a5+3,column=i+1).value = ws1.cell(row=2,column=i+1).value
      ws1.cell(row=a1+5+a2+3+a3+3+a4+3+a5+3,column=i+1).font=Font(bold=True)

    ws1.insert_rows(a1+5+a2+3+a3+3+a4+3+a5+3+a6+1)
    ws1.insert_rows(a1+5+a2+3+a3+3+a4+3+a5+3+a6+2)
    ws1['A'+str(a1+5+a2+3+a3+3+a4+3+a5+3+a6+2)]='Work Not Yet Due'
    ws1.insert_rows(a1+5+a2+3+a3+3+a4+3+a5+3+a6+3)

    for i in range(ws1.max_column):
      ws1.cell(row=a1+5+a2+3+a3+3+a4+3+a5+3+a6+3,column=i+1).value = ws1.cell(row=2,column=i+1).value
      ws1.cell(row=a1+5+a2+3+a3+3+a4+3+a5+3+a6+3,column=i+1).font=Font(bold=True)

    ws1.insert_rows(a1+5+a2+3+a3+3+a4+3+a5+3+a6+3+a7+1)
    ws1.insert_rows(a1+5+a2+3+a3+3+a4+3+a5+3+a6+3+a7+2)
    ws1['A'+str(a1+5+a2+3+a3+3+a4+3+a5+3+a6+3+a7+2)]='In 2 Strike Rule'
    ws1.insert_rows(a1+5+a2+3+a3+3+a4+3+a5+3+a6+3+a7+3)

    for i in range(ws1.max_column):
      ws1.cell(row=a1+5+a2+3+a3+3+a4+3+a5+3+a6+3+a7+3,column=i+1).value = ws1.cell(row=2,column=i+1).value
      ws1.cell(row=a1+5+a2+3+a3+3+a4+3+a5+3+a6+3+a7+3,column=i+1).font=Font(bold=True)


    wb1.save(filename)
    myorder=[0,1,2,3,4,5,6,7,9,8]
    wb1._sheets =[wb1._sheets[s] for s in myorder]
    wb1.save(filename)    
